#Server socket python
# Importing stuff
import os
import socket
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook
from captcha.image import ImageCaptcha
#ACG imports
from hashlib import pbkdf2_hmac, sha256
from Cryptodome.Random import get_random_bytes
from Cryptodome.Cipher import PKCS1_OAEP, PKCS1_v1_5, AES
from Cryptodome.Util.Padding import pad, unpad
from Cryptodome.PublicKey import RSA
from Cryptodome.Signature import pkcs1_15
from Cryptodome.Hash import SHA256
import base64

# Get today's day
today = datetime.today()
today_day = today.isoweekday()
weekdays = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

# Loading a excel file in Python
wb = load_workbook(filename='AssignmentData.xlsx')
membership_sheet = wb['Membership'] # Set membership_sheet to Membership sheet
customers_sheet = wb['Customers'] # Set customers_seet to Customers sheet
preorder_sheet = wb['Preorder'] # Set membership_sheet to Membership sheet


# Login functions
class Login:
    def __init__(self):
        self.membership = {}

        # Creating dictionary of username : salted password
        for row in membership_sheet.iter_rows(min_row=2, max_col=3, max_row=membership_sheet.max_row, values_only=True):
            self.membership[row[1]] = row[2]

    # Function to check if user is in filesystem
    def user_in_system(self, user):
        if user in self.membership:
            return True
        else:
            return False

    # Function to add a registered user to filesystem
    def add_new_user(self, user, hash):
        membership_sheet.append([len(self.membership)+1, user, hash])
        self.membership[user] = hash

    # Function to retrieve salt of user and return it
    def send_salt(self, user):
        return self.membership[user][:64]

    # Function to check if user credential matches  
    def verify_login(self, user, hash):
        if hash == self.membership[user][64:]:
            return True
        else:
            return False

    #  Function to generate a random captcha image
    def captcha(self):
        image = ImageCaptcha()
        self.random_captcha = os.urandom(2).hex()
        data = image.generate(self.random_captcha)
        return data.getvalue()

    # Function to validate if captcha was correctly inputed
    def captcha_validation(self, raw_captcha):
        if raw_captcha == self.random_captcha:
            return True
        else:
            return False


# FileSystem's functions
class FileSystem:
    try:
        # Function to retrieve a selected day food menu
        def process_foodmenu(self, day=today_day):
            sheet = wb[weekdays[int(day)-1]] # Set sheet to current day of the week

            # Creating string of "menu:price,discount |" based on the current day of the week
            food_menus = ""
            for row in sheet.iter_rows(min_row=2, max_col=3, max_row=sheet.max_row, values_only=True):
                food_menus += f"{row[0]}:{str(row[1])},{str(row[2])}|"

            return food_menus

        # Function to add a new menu item on a given day
        def add_new_menu(self, day, foodname, price, discount):
            sheet = wb[weekdays[int(day)-1]] # Set sheet to current day of the week
            sheet.append([foodname.strip(), float(price.strip()), float(discount.strip())/100])

        # Function to remove a menu item on a given day
        def remove_menu(self, day, row):
            sheet = wb[weekdays[int(day)-1]] # Set sheet to current day of the week
            sheet.delete_rows(int(row)+1, 1)

        # Function to edit a menu item's price/discount on a given day
        def edit_menu(self, day, row, price, discount):
            sheet = wb[weekdays[int(day)-1]] # Set sheet to current day of the week
            sheet["B"+str(int(row)+1)], sheet["C"+str(int(row)+1)] = float(price), float(discount)/100 # at current day sheet cell B(whichever number of choice + 1) to be set price and cell C(whichever number of choice + 1) to be set discount

        # Function to append the user's preordered food
        def preorder(self, user, food_preordered):
            preorder_sheet.append([user, food_preordered])

        # Function to check a user's preorder if any.
        def check_preorder(self, user):
            for row_number, row in enumerate(preorder_sheet.iter_rows(min_row=2, max_col=2, max_row=preorder_sheet.max_row, values_only=True), 2):
                if row[0] == user and weekdays[today_day-1] in row[1]:
                    preorder = json.loads(row[1])
                    preorder_to_send = preorder[weekdays[today_day-1]]

                    del preorder[weekdays[today_day-1]]
                    preorder_sheet.delete_rows(row_number)
                    if len(preorder) > 0:
                        preorder_sheet.insert_rows(row_number)
                        preorder_sheet["A"+str(row_number)], preorder_sheet["B"+str(row_number)] = user, json.dumps(preorder)

                    return json.dumps(preorder_to_send)
            else:
                return "NULL"

        def record_transaction(self, user, food_ordered):
            customers_sheet.append([today.strftime("%c"), user, food_ordered])

    finally:
        # Saves the excel file
        wb.save('AssignmentData.xlsx')


def start_server():#This starts the server and waits for response from function "handler"
    soc = socket.socket(socket.AF_INET, socket.SOCK_STREAM)

    try:
        soc.bind(('127.0.0.1', 8089)) # Binds localhost and port 8089 to socket
    except socket.error as msg:
        import sys
        print('Bind failed. Error : ' + str(sys.exc_info()))
        print(msg.with_traceback())
        sys.exit() # Exit if any error
    soc.listen(1) # Listen for 1 connections
    while True:
        print("waiting a new call at accept()")
        global con
        con, address = soc.accept() # Accept a connection
        if handler(con) == 'SHUTDOWN':
            break
    soc.close()
    print(f"Server stopped at {today}")

def handler(con):#This handles server connections input
    print("client connected")
    while True:
        buf = con.recv(8192) # buf is of the type of byte
        content = buf.decode() # decode buf into a string
        if len(buf) > 0:
            ##Step 1 (GETS CLIENT PUBLIC KEY AND RETURN SERVER PUBLIC KEY)
            if content.startswith("CLIENTPUBLICKEY"):#Client request new RSA key
                global cryptothingy
                client_public_key=content.split("$")[1].strip()
                cryptothingy=Cryptostuff(client_public_key)
                continue#Reloop handler once public key is gained
            ##Step 2 (Gets session key from Client [will reloop if invalid])
            if cryptothingy.aes_session_cipher=="NULL":#This waits untill a session key is created
                cryptothingy.get_session_key(content)#session key will not be changed hence will reloop if session key invalid
                continue#reloop handler to wait for new message
            else:
            ##Step 3 (Decrypts message so that it is readable)
                content = base64.b64decode(buf)
                content=cryptothingy.aes_decrypt(content)#decrypts request
                print("Session key used")
                content = content.decode()
                print(f"Request from Client: {content}")
            ##Step 4 (Check client's request)
            if content == 'SHUTDOWN':
                return content
            elif content.startswith("CHECKUSER"):
                if login.user_in_system(content.split()[1].strip()):
                    encrypted_message_to_send=cryptothingy.aes_encrypt('TRUE')
                    con.sendall(encrypted_message_to_send)
                else:
                    encrypted_message_to_send=cryptothingy.aes_encrypt('FALSE')
                    con.sendall(encrypted_message_to_send)

            elif content.startswith("ADDUSER"):
                user = content.split()[1].strip()
                hash = content.split()[2].strip()
                login.add_new_user(user, hash)

            elif content == "CAPTCHA":#This is not encrypted
                con.sendall(login.captcha())#ONLY THIS IS NOT ENCRYPTED

            elif content.startswith("CAPTCHAVALIDATION"):
                raw_captcha = content.split()[1].strip()
                if login.captcha_validation(raw_captcha):
                    encrypted_message_to_send=cryptothingy.aes_encrypt('TRUE')
                    con.sendall(encrypted_message_to_send)
                else:
                    encrypted_message_to_send=cryptothingy.aes_encrypt('FALSE')
                    con.sendall(encrypted_message_to_send)

            elif content.startswith("SENDSALT"):
                encrypted_message_to_send=cryptothingy.aes_encrypt(login.send_salt(content.split()[1]))
                con.sendall(encrypted_message_to_send)

            elif content.startswith("VERIFYLOGIN"):
                user = content.split()[1].strip()
                hash = content.split()[2].strip()
                if login.verify_login(user, hash):
                    encrypted_message_to_send=cryptothingy.aes_encrypt('TRUE')
                    con.sendall(encrypted_message_to_send)
                else:
                    encrypted_message_to_send=cryptothingy.aes_encrypt('FALSE')
                    con.sendall(encrypted_message_to_send)

            elif content.startswith("CHECKPREORDER"):
                user = content.split()[1].strip()
                message_to_send=FileSystem().check_preorder(user)
                encrypted_message_to_send=cryptothingy.aes_encrypt(message_to_send)
                con.sendall(encrypted_message_to_send)

            elif content.startswith("RETRIVEMENU") and content[-1].isnumeric():
                message_to_send=FileSystem().process_foodmenu(content[-1].strip())
                message_signature=cryptothingy.generate_digitalsignature(message_to_send)
                send_to_client=message_to_send.encode()+b"$"+base64.b64encode(message_signature)
                encrypted_message_to_send=cryptothingy.aes_encrypt(send_to_client.decode())
                con.sendall(encrypted_message_to_send)

            elif content.startswith("ADDMENU"):
                day = content.split('|')[1].strip()
                foodname = content.split('|')[2].strip()
                price = content.split('|')[3].strip()
                discount = content.split('|')[4].strip()
                FileSystem().add_new_menu(day, foodname, price, discount)

            elif content.startswith("REMOVEMENU") and content[-1].isnumeric():
                day = content.split()[1].strip()
                row = content.split()[2].strip()
                FileSystem().remove_menu(day, row)

            elif content.startswith("EDITMENU"):
                day = content.split()[1].strip()
                row = content.split()[2].strip()
                price = content.split()[3].strip()
                discount = content.split()[4].strip()
                FileSystem().edit_menu(day, row, price, discount)

            elif content.startswith("PREORDER"):
                user = content.split('|')[1].strip()
                preorder = content.split('|')[2].strip()
                FileSystem().preorder(user, preorder)

            elif content.startswith("RECORDTRANSACTION"):
                valid_bool=cryptothingy.check_digitalsignature(content)
                if valid_bool:
                    content=content.split("$")[0].strip()
                    user = content.split('|')[1].strip()
                    food_ordered = content.split('|')[2].strip()
                    FileSystem().record_transaction(user, food_ordered)
                    print("New transaction log added.")
                else:
                    print("Invalid Digital Signature. Record Transaction request ingnored.")
            cryptothingy.aes_session_cipher="NULL"
            print("AES Session cipher erased")
        else: # 0 length buf implies client has dropped the con.
            print("client disconnected")
            break # quit this handler immediately and return ""
    # Saves the excel file
    wb.save('AssignmentData.xlsx')
    con.close() #exit from the loop when client sent q or x
    return buf.decode()

#Additional Codes for ACG
class Cryptostuff:
    def __init__(self,client_public_key):#This function generates the RSA key.
        self.client_public_key=client_public_key
        print("Generating RSA Key on server side...")
        self.rsa_keypair=RSA.generate(2048)
        self.server_private_key=self.rsa_keypair.exportKey().decode()
        self.server_public_key=self.rsa_keypair.publickey().exportKey().decode()
        self.aes_session_cipher="NULL"
        print("AES Session cipher erased")
        con.sendall(self.server_public_key.encode())
        print("Public key sent to client")
        print("Key exchange success on server side.")
        return

    def rsa_decryption(self,encrypted_message):#handles rsa decryption
        print("RSA decryption...")
        server_private_rsa_cipher = PKCS1_OAEP.new(RSA.import_key(self.server_private_key.encode()))#Client's public key
        decrypted_message = server_private_rsa_cipher.decrypt(encrypted_message)
        print("RSA decryption success")
        return decrypted_message

    def get_session_key(self,content):#handles storing session key
        try:#incase something goes wrong in split
            if content.startswith("1$"):
                print("Get Session Key...")
                encrypted_AES_key_WITH_RSA=content.split("$")[1]#Encrypted AES Key
                session_iv=content.split("$")[2]
                encrypted_AES_key_WITH_RSA = bytes.fromhex(encrypted_AES_key_WITH_RSA)
                self.session_iv = bytes.fromhex(session_iv)
                self.aes_session_key=self.rsa_decryption(encrypted_AES_key_WITH_RSA)
                self.aes_session_cipher = AES.new(self.aes_session_key,AES.MODE_CBC,iv=self.session_iv)#Stores session key to class
                print("Session Key successfully grabbed")
                con.sendall("1$".encode())
            else:
                print("Invalid session key")
                con.sendall("0$".encode())
        except:
            con.sendall("0$".encode())
        return

    def aes_decrypt(self,content):#handles aes decryption
        print("Aes Decrypting...")
        plain_text=unpad(self.aes_session_cipher.decrypt(content),AES.block_size)
        print("Aes Decryption success")
        return plain_text
    
    def aes_encrypt(self,content):#handles aes encryption
        self.aes_session_cipher = AES.new(self.aes_session_key,AES.MODE_CBC,iv=self.session_iv)#Stores session key to class
        print("AES encrypting...")
        ciphertext = self.aes_session_cipher.encrypt(pad(content.encode(), AES.block_size))
        print("AES succesfully encrypted")
        return ciphertext

    def generate_digitalsignature(self,content):#Returns digital signature for server with content
        print("Generating Digital Signature")
        digest = SHA256.new(content.encode())
        signer = pkcs1_15.new(self.rsa_keypair)
        signature = signer.sign(digest)
        print("Digital Signature created.")
        return signature
    
    def check_digitalsignature(self,content):#Returns whether digital signature is valid
        print("Checking Client's Digital Signature")
        message = content.split("$")[0]
        signature = content.split("$")[1]
        signature = base64.b64decode(signature)
        digest=SHA256.new(message.encode())
        verifier = pkcs1_15.new(RSA.import_key(self.client_public_key.encode()))
        try:
            verifier.verify(digest,signature)
            print("The signature is valid...")
            return True
        except:
            print("The signature is not valid...")
            return False
    
# Start program
login = Login()
start_server()